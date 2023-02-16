import json
import pandas as pd
import datetime
import os
import openpyxl
from openpyxl.styles import Border, Side
import shutil
import time

#Time
current_month = datetime.datetime.now().strftime("%B")
current_year = int(datetime.datetime.now().strftime("%Y")) + 543


with open('/volume1/web/ComputerLab-SRV-Checkin/form.json', 'rb') as file:
    formjson = file.read().decode('utf-8')
    formjson = json.loads(formjson)

with open('/volume1/web/ComputerLab-SRV-Checkin/form.json', 'w', encoding='utf-8') as file:
    json.dump(formjson, file, ensure_ascii=False)

df = pd.DataFrame(formjson)
# UTF-8 Decode
with open(f'/volume1/web/ComputerLab-SRV-Checkin/formDataJson/{current_month}_{current_year}.json', 'w', encoding='utf-8') as file:
    json.dump(formjson, file, ensure_ascii=False)


# Save for user

month_thai = {
"January": "มกราคม",
"February": "กุมภาพันธ์",
"March": "มีนาคม",
"April": "เมษายน",
"May": "พฤษภาคม",
"June": "มิถุนายน",
"July": "กรกฎาคม",
"August": "สิงหาคม",
"September": "กันยายน",
"October": "ตุลาคม",
"November": "พฤศจิกายน",
"December": "ธันวาคม"
}

current_month_thai = month_thai[current_month]

file_name = f"/volume1/web/ComputerLab-SRV-Checkin/formData/{current_month}_{current_year}.xlsx"

# Create a list of headers
headers = ['ที่', 'วัน-เดือน-ปี', 'รหัสนักเรียน', 'ชื่อ', 'นามสกุล', 'ชั้น', 'เพศ', 'เข้า', 'ออก', 'Lab Room', 'หมายเหตุ']

# Write the headers and the title to the worksheet
with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
    df.to_excel(writer, index=False, header=False, startrow=5)
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']
    for i, header in enumerate(headers):
        worksheet.cell(row=5, column=i+1, value=header)
        worksheet.cell(row=5, column=i+1).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
        worksheet.cell(row=5, column=i+1).border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                                            top=Side(style='thin'), bottom=Side(style='thin'))
    worksheet.cell(row=1, column=5, value='แบบบันทึกการขอใช้ห้องปฏิบัติการคอมพิวเตอร์')
    worksheet.cell(row=1, column=5).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    worksheet.cell(row=2, column=5, value=f'ประจำเดือน {current_month_thai} พ.ศ.{current_year}')
    worksheet.cell(row=2, column=5).alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    worksheet.row_dimensions[4].height = 20
    worksheet.row_dimensions[5].height = 30
    worksheet.column_dimensions[worksheet.cell(row=5, column=4).column_letter].width = 20
    worksheet.column_dimensions[worksheet.cell(row=5, column=5).column_letter].width = 20